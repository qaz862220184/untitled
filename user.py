import asyncio
import pyppeteer
from lxml import etree
from pyppeteer import launch,launcher
import random

import openpyxl as xl
import os
import time
def screen_size():
    # 使用tkinter获取屏幕大小
    import tkinter
    tk = tkinter.Tk()
    width = tk.winfo_screenwidth()
    height = tk.winfo_screenheight()
    tk.quit()
    return width, height


url1 = 'https://www.fakeaddressgenerator.com/'

async def get_user_detail():
    user_detail={}
    browser = await launch({
        'executablePath': pyppeteer.launcher.executablePath(),
        #'headless': False,
        'dumpio': True,
        'autoClose': False,
        'args': [
            '--no-sandbox',
            "--start-maximized",
            '--disable-infobars',
            '--disable-extensions',
            '--hide-scrollbars',
            '--disable-bundled-ppapi-flash',
            '--mute-audio',
            #'--no-sandbox',
            '--disable-setuid-sandbox',
            '--disable-gpu',
            #'--proxy-server={}'.format(proxy_ip)
            #'--proxy-server=192.168.1.250:7890',
            '--proxy-server=192.168.1.239:24004'
        ]
    })

    page = await browser.newPage()
    width, height = screen_size()
    # print(width,height)
    await page.setViewport({  # 最大化窗口
        "width": width,
        "height": 900
    })
    await page.evaluateOnNewDocument('() =>{ Object.defineProperties(navigator,'
                                     '{ webdriver:{ get: () => false } }) }')
    await page.goto(url1)
    await page.waitFor(5000)
    contents = await page.content()
    tree = etree.HTML(contents)
    tr_list = tree.xpath('/html/body/div[1]/div[3]/div[1]/div/div/div[2]/div[1]/div[2]/table/tbody/tr')

    for tr in tr_list:
        a = tr.xpath('.//td[1]//text()')
        a = a[0]
        b = tr.xpath('.//td[2]//text()')
        b = b[0]
        user_detail.update({a: b})

    div_list = tree.xpath('//*[@class="row item"]')
    for i in range(7):
        a = div_list[i].xpath('.//div[1]//text()')
        a = a[0]
        b = div_list[i].xpath('.//div[2]//input/@value')[0]
        user_detail.update({a: b})

    # 获取email地址

    await page.waitFor(2000)
    page1 = await browser.newPage()
    await page1.setViewport({  # 最大化窗口
        "width": width,
        "height": 900
    })
    await asyncio.wait([
        page.goto('https://email-fake.com'),
        page.waitFor('#email_ch_text'),
    ])
    email = await page.Jeval('#email_ch_text', 'el => el.innerHTML')
    print(email)

    user_detail.update({'email': email})
    print(user_detail)
    await browser.close()
    return user_detail


def write_excel_file(folder_path,user_detail):
    result_path = os.path.join(folder_path, "user_detail.xlsx")
    print(result_path)
    detail = []
    print('***** 开始写入excel文件 ' + result_path + ' ***** \n')
    if os.path.exists(result_path):
        print('***** excel已存在，在表后添加数据 ' + result_path + ' ***** \n')
        workbook = xl.load_workbook(result_path)
        sheet = workbook.active
        for key in user_detail:
            detail.append(user_detail[key])
        sheet.append(detail)
        workbook.save(result_path)
    else:
        print('***** excel不存在，创建excel ' + result_path + ' ***** \n')
        workbook = xl.Workbook()
        workbook.save(result_path)

        sheet = workbook.active
        for item in user_detail:
            headers.append(item)
        sheet.append(headers)
        for key in user_detail:
            detail.append(user_detail[key])

        sheet.append(detail)
        workbook.save(result_path)
    print('***** 生成Excel文件 ' + result_path + ' ***** \n')



def write_ui_qt(folder_path,user_detail):
    now = time.strftime("%d_%m_%Y")
    result_path = os.path.join(folder_path, "qt_%s.xlsx" %now)
    print(result_path)
    headers = ['Full Name','email', 'tag']
    register_detail = []
    register_detail.append(user_detail['Full Name'])
    register_detail.append(user_detail['email'])
    register_detail.append('申请中')
    print('***** 开始写入excel文件 ' + result_path + ' ***** \n')
    if os.path.exists(result_path):
        print('***** excel已存在，在表后添加数据 ' + result_path + ' ***** \n')
        workbook = xl.load_workbook(result_path)
        sheet = workbook.active
        sheet.append(register_detail)
        workbook.save(result_path)
    else:
        print('***** excel不存在，创建excel ' + result_path + ' ***** \n')
        workbook = xl.Workbook()
        workbook.save(result_path)

        sheet = workbook.active
        sheet.append(headers)

        sheet.append(register_detail)
        workbook.save(result_path)
    print('***** 生成Excel文件 ' + result_path + ' ***** \n')

def main():

    global detail
    detail = []
    global headers
    headers = []
    global user_detail
    user_detail = {}
    print('是否进入到main函数')
    try:
        asyncio.get_event_loop().run_until_complete(get_user_detail())
        print('是否执行了获取函数')

        #write_excel_file("./")
        #write_ui_qt('./')
    except Exception as e:
        print(e)
        print('获取用户信息失败')

    #return user_detail








if __name__ == '__main__':
    while True:
        try:
            user_detail = asyncio.get_event_loop().run_until_complete(get_user_detail())
            write_excel_file("./",user_detail)
            write_ui_qt('./',user_detail)
        except Exception:
            print('获取用户信息失败')
            continue
            


