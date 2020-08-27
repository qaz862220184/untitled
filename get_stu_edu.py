import asyncio
import pyppeteer
import openpyxl as xl
import os
from lxml import etree
from pyppeteer import launch
import re
import time
#email = 'qifcz5e6@temporary-mail.net'



def screen_size():
    # 使用tkinter获取屏幕大小
    import tkinter
    tk = tkinter.Tk()
    width = tk.winfo_screenwidth()
    height = tk.winfo_screenheight()
    tk.quit()
    return width, height



async def get_edu(email_detail):

    browser = await launch({
        'executablePath': pyppeteer.launcher.executablePath(),
        #'headless': False,
        'dumpio': True,
        'autoClose': True,
        'args': [
            '--no-sandbox',
            "--start-maximized",
            '--disable-infobars',
            '--disable-extensions',
            '--hide-scrollbars',
            '--disable-bundled-ppapi-flash',
            '--mute-audio',
            '--no-sandbox',
            '--disable-setuid-sandbox',
            '--disable-gpu',
            '--enable-automation',
            #'--proxy-server={}'.format(proxy_ip)
            #'--proxy-server=192.168.1.250:7890',
            '--proxy-server=192.168.1.239:24003'
        ]
    })
    page = await browser.newPage()
    width, height = screen_size()
    # print(width,height)
    await page.setViewport({  # 最大化窗口
        "width": width,
        "height": 900
    })
    await page.evaluateOnNewDocument('() =>{ Object.defineProperties(navigator,'
                                     '{ webdriver:{ get: () => false } }) }')


    '''
    # 获取email地址
    url1 = 'https://www.temporary-mail.net/mailbox/'+ email_detail
    #await page.goto('https://www.temporary-mail.net/change')
    await page.goto(url1)
    await page.waitFor(10000)

    email_href = ''
    contents = await page.content()
    tree = etree.HTML(contents)
    tr_list = tree.xpath('//*[@id="message-list"]/tr')
    print(tr_list)
    for tr in tr_list:
        email_title = tr.xpath('./td[2]//text()')[0]
        #email_sender = tr.xpath('./td[1]/a/text()')[0]
        #print(email_title,email_sender)
        #sender = 'NO_Reply'
        if email_title == 'WELCOME TO MERCED COLLEGE':
            print('就是这封邮件')
            email_href = tr.xpath('./td[2]/a/@href')[0]
            break

    if email_href =='':
        return 0
    else:
        email_url = 'https://www.temporary-mail.net' + email_href
        print(email_url)
        await page.goto(email_url)
        contents1 = await page.content()
        tree1 = etree.HTML(contents1)
        text_part = ''
        detail = tree1.xpath('//*[@id="main"]/div/div/div[2]/div[1]/div/div[3]/text()')
        for data in detail:
            text_part += data

        stu_id = re.search('\d{7}',text_part).group()
        print(stu_id)

        # edu_email = re.findall(r'([\w-]+(\.[\w-]+)*@[\w-]+(\.[\w-]+)+)',text_part)
        # print(edu_email)

        edu_email = re.search(r'([\w-]+(\.[\w-]+)*(@campus\.mccd\.edu))', text_part).group()
        print(edu_email)
        #print(text_part)
        '''
    url = 'https://email-fake.com/' + email_detail
    print(url)

    await page.goto(url)

    content = await page.content()
    tree = etree.HTML(content)
    email_href = ''
    em_list = tree.xpath('//*[@id="email-table"]/a')
    for em in em_list:
        title = em.xpath('./div[2]/text()')
        if title[0] == 'WELCOME TO MERCED COLLEGE':
            print('就是这封邮件')
            email_href = em.xpath('./@href')[0]
            break
    if email_href == '':
        return 0
    else:
        email_url = 'https://email-fake.com' + email_href
        print(email_url)
        await page.goto(email_url)
        contents1 = await page.content()
        tree1 = etree.HTML(contents1)
        text_part = ''
        # detail = tree1.xpath('//*[@id="main"]/div/div/div[2]/div[1]/div/div[3]/text()')
        detail = tree1.xpath('//*[@id="iddelet2"]/div[4]/div[3]/p/text()')
        for data in detail:
            text_part += data
        stu_id = re.search('\d{7}', text_part).group()
        print(stu_id)
        edu_email = tree1.xpath('//*[@id="iddelet2"]/div[4]/div[3]/p/a[4]/text()')[0]
        print(edu_email)
        return stu_id,edu_email





        ## 接下来就是从detail中取到学生id和邮箱地址

def write_excel_file(folder_path,stu_id,edu_email,edu_pwd):
    result_path = os.path.join(folder_path, "stu_edu.xlsx")
    print(result_path)
    headers = ['stu_id', 'edu_email', 'edu_pwd']
    stu_edu = []
    stu_edu.append(stu_id)
    stu_edu.append(edu_email)
    stu_edu.append(edu_pwd)
    print('***** 开始写入excel文件 ' + result_path + ' ***** \n')
    if os.path.exists(result_path):
        print('***** excel已存在，在表后添加数据 ' + result_path + ' ***** \n')
        workbook = xl.load_workbook(result_path)
        sheet = workbook.active
        sheet.append(stu_edu)
        workbook.save(result_path)
    else:
        print('***** excel不存在，创建excel ' + result_path + ' ***** \n')
        workbook = xl.Workbook()
        workbook.save(result_path)

        sheet = workbook.active
        sheet.append(headers)

        sheet.append(stu_edu)
        workbook.save(result_path)
    print('***** 生成Excel文件 ' + result_path + ' ***** \n')

def red_excel_file(folder_path):
    stu_register = {}
    result_path = os.path.join(folder_path, "stu_register.xlsx")
    wb2 = xl.load_workbook(result_path)
    sheet = wb2.active

    cell = sheet[1]
    cell1 = sheet[2]
    for i in range(len(cell)):
        stu_register.update({cell[i].value: cell1[i].value})

    return stu_register


def remove_excel_file(folder_path):
    result_path = os.path.join(folder_path, "stu_register.xlsx")
    wb2 = xl.load_workbook(result_path)
    sheet = wb2.active
    sheet.delete_rows(2)
    wb2.save(result_path)


def main():


    stu_register = red_excel_file('./')
    email = stu_register['email']
    Birthday = stu_register['Birthday']
    print(email, Birthday)
    global stu_id,edu_email,edu_pwd
    stu_id, edu_email = '', ''
    m, d, y = Birthday.split('/')
    m = m.zfill(2)
    d = d.zfill(2)
    y = y[-2:]
    edu_pwd = m + d + y
    try:
        print('调用前')
        #stu_id, edu_email = get_edu(email)
        stu_id, edu_email = asyncio.get_event_loop().run_until_complete(get_edu(email))
        print('调用后')
    except Exception as e:
        print(e)
        print('获取邮箱信息失败')
    print(stu_id, edu_email)
    if stu_id and edu_email and edu_pwd:
        write_excel_file('./',stu_id,edu_email,edu_pwd)
        remove_excel_file('./')
    else:
        print('未获得完整学生邮箱信息')

if __name__ == '__main__':

    while True:
        try:
            main()
        except Exception as e:
            print(e)
            time.sleep(300)
            continue

'''
if __name__ == '__main__':
    stu_register = {}
    headers = ['stu_id','edu_email','edu_pwd']
    red_excel_file('./')
    print(stu_register)

    email=stu_register['email']
    Birthday = stu_register['Birthday']
    print(email,Birthday)
    stu_id, edu_email = '',''
    m, d, y = Birthday.split('/')
    m = m.zfill(2)
    d = d.zfill(2)
    y = y[-2:]
    edu_pwd = m + d + y
    try:
        stu_id, edu_email = asyncio.get_event_loop().run_until_complete(get_edu(email))
    except Exception:
        print('获取邮箱信息失败')
    print(stu_id,edu_email)
    if stu_id and edu_email and edu_pwd:
        write_excel_file('./')
        remove_excel_file('./')
    else:
        print('未获得完整学生邮箱信息')
'''

